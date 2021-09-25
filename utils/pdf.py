import locale
from datetime import datetime
from openpyxl import load_workbook

locale.setlocale(locale.LC_ALL, "fr_FR")

month_start_col = "d"
month_start_line = "3"
month_offset_col = 3
day_offset_line = 3
event_offset_line = 2


def detect_month_column(ws):
    month = datetime.now().strftime("%B").upper()
    index = f"{month_start_col}{month_start_line}"
    for i in range(0, 11):
        if month == ws[index].value:
            return index[0]
        index = f"{chr(ord(month_start_col) + i + month_offset_col)}{month_start_line}"


def detect_day_line():
    return datetime.now().day + day_offset_line


def detect_calendar(ws, month_pos, day_pos, max_days_to_check):
    events = []
    i = 0
    event = {}
    actual_color = ""
    for _ in range(0, max_days_to_check + 1):
        # At end of month, month_pos++
        if not isinstance(ws[f"{month_pos}{day_pos + i}"].value, int):
            day_pos = day_offset_line + 1
            month_pos = f"{chr(ord(month_pos) + month_offset_col)}"
            i = 0
        index_event = f"{chr(ord(month_pos) + event_offset_line)}{day_pos + i}"
        index_month = f"{month_pos}{month_start_line}"
        # Event is not finished
        if ws[index_event].fill.start_color.index == actual_color:
            # If there is an information in the event cell
            if ws[index_event].value is not None:
                event["title"] = f"{event['title']} {ws[index_event].value}"
            event[
                "end"
            ] = f"{ws[f'{month_pos}{day_pos + i}'].value} {ws[index_month].value.title()}"
        # If color change maybe an event is detected
        if ws[index_event].fill.start_color.index != actual_color:
            # Add the event if he has a title (don't add weekend as event)
            if "title" in event and event["title"] != "":
                events.append(event)
            event = {}
            actual_color = ws[index_event].fill.start_color.index
            # If there is an information in the event cell
            if ws[index_event].value:
                event["title"] = ws[index_event].value
            else:
                event["title"] = ""
            event[
                "start"
            ] = f"{ws[f'{month_pos}{day_pos + i}'].value} {ws[index_month].value.title()}"
            event[
                "end"
            ] = f"{ws[f'{month_pos}{day_pos + i}'].value} {ws[index_month].value.title()}"

        i = i + 1
    return events


def get_events(pdf, max_days_to_check):
    wb = load_workbook(filename=pdf)
    ws = wb.active

    month_pos = detect_month_column(ws)
    day_pos = detect_day_line()
    events = detect_calendar(
        ws,
        month_pos,
        day_pos,
        max_days_to_check,
    )

    return events
